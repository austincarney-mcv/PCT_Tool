#!/usr/bin/env python3
"""
Seed PCT demo project from Excel into SQLite.
Source: 10162 - Project Control Tool.xlsx
Target: data/pct.db
"""

import sqlite3
import openpyxl
import sys
from datetime import datetime

DB = 'data/pct.db'
XL = '10162 - Project Control Tool.xlsx'

# ─── helpers ────────────────────────────────────────────────────────────────

def d(v):
    """Parse a date value to YYYY-MM-DD string, or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        s = v.strip().split('\n')[0].strip()
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
    return None

def f(v, default=0.0):
    """Cast to float, return default on failure."""
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def tx(v):
    """Stringify and normalise whitespace/non-breaking spaces."""
    if v is None:
        return None
    r = str(v).replace('\xa0', ' ').replace('\n', ' ').strip()
    r = ' '.join(r.split())
    return r if r else None

def norm_disc(v):
    """Map Excel section-header discipline label to DB enum value.
    Only matches exact section headers, not arbitrary strings containing keywords."""
    v = str(v).strip().upper()
    # Strip ' RESOURCE' or ' RESOURCE PROGRAM' suffix (exact, not character-strip)
    for sfx in (' RESOURCE PROGRAM', ' RESOURCE'):
        if v.endswith(sfx):
            v = v[:-len(sfx)]
            break
    mapping = {
        'ARCHITECTURE':    'Architecture',
        'ARCHITECTURAL':   'Architecture',
        'CIVIL':           'Civil',
        'STRUCTURES':      'Structural',
        'STRUCTURAL':      'Structural',
        'HYDRAULICS':      'Hydraulics',
        'LANDSCAPING':     'Landscaping',
        'CERTIFIER':       'Certifier',
        'FIRE ENGINEERING':'Fire Engineering',
        'FIRE SERVICES':   'Fire Services',
        'BUILDER/CM':      'Builder/CM',
        'BUILDER':         'Builder/CM',
    }
    return mapping.get(v)

def rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


# ─── main ────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(XL, data_only=True)
    con = sqlite3.connect(DB)
    con.execute('PRAGMA foreign_keys = ON')
    cur = con.cursor()

    # ── 1. PROJECT ──────────────────────────────────────────────────────────
    print('Inserting project…')
    cur.execute('SELECT id FROM projects WHERE project_number = ?', ('10162',))
    row = cur.fetchone()
    if row:
        pid = row[0]
        cur.execute(
            '''UPDATE projects SET project_name=?, client=?, author=?,
               date_created=?, version=?, release_status=?, updated_at=datetime('now')
               WHERE id=?''',
            ('Lot 31 The Hub Heathwood', 'Prekaro Projects', 'Stephen Davies',
             '2025-09-08', '1.0', 'Draft', pid)
        )
        print(f'  Updated existing project id={pid}')
    else:
        cur.execute(
            '''INSERT INTO projects (project_number, project_name, client, author,
               date_created, version, release_status)
               VALUES (?,?,?,?,?,?,?)''',
            ('10162', 'Lot 31 The Hub Heathwood', 'Prekaro Projects',
             'Stephen Davies', '2025-09-08', '1.0', 'Draft')
        )
        pid = cur.lastrowid
        print(f'  Created project id={pid}')
    # Clear all child tables first (FK-safe order)
    cur.execute('DELETE FROM c2c_resource_allocations WHERE snapshot_id IN '
                '(SELECT id FROM c2c_snapshots WHERE project_id=?)', (pid,))
    cur.execute('DELETE FROM c2c_discipline_financials WHERE snapshot_id IN '
                '(SELECT id FROM c2c_snapshots WHERE project_id=?)', (pid,))
    cur.execute('DELETE FROM c2c_snapshots WHERE project_id=?', (pid,))
    for t in ('team_resources', 'drawings', 'approvals', 'critical_items',
              'design_changes', 'risks', 'rfis', 'lessons_learnt', 'sid_hazards',
              'value_log', 'brief_compliance'):
        cur.execute(f'DELETE FROM {t} WHERE project_id=?', (pid,))
    con.commit()

    # ── 2. TEAM RESOURCES ───────────────────────────────────────────────────
    print('Inserting team resources...')
    resources_design = [
        # Architecture
        ('Michael McVeigh',            488, 'Architecture', 1),
        ('Elliot Blucher',             284, 'Architecture', 2),
        ('Sonya Butt',                 244, 'Architecture', 3),
        ('Architectural Technician',   192, 'Architecture', 4),
        ('Architectural Graduate',     140, 'Architecture', 5),
        ('Asst PM Shannon',            192, 'Architecture', 6),
        # Civil
        ('Michael McVeigh',            488, 'Civil', 1),
        ('Winnie Wu',                  284, 'Civil', 2),
        ('Carlos Sailema Cobo',        192, 'Civil', 3),
        ('Steph Bennett',              244, 'Civil', 4),
        ('Jerome Bauyon',              140, 'Civil', 5),
        ('Asst PM Shannon',            192, 'Civil', 6),
        # Structural
        ('Michael McVeigh',            488, 'Structural', 1),
        ('Stephen Davies',             376, 'Structural', 2),
        ('Atef Raouf',                 328, 'Structural', 3),
        ('Jack Pham',                  164, 'Structural', 4),
        ('Jason McGregor',             192, 'Structural', 5),
        ('Asst PM Shannon',            192, 'Structural', 6),
    ]
    # Remove existing resources for project to avoid duplicates on re-run
    cur.execute('DELETE FROM team_resources WHERE project_id=?', (pid,))
    resource_id_map = {}  # (name, discipline) -> id
    for name, rate, disc, sort in resources_design:
        cur.execute(
            'INSERT INTO team_resources (project_id, name, discipline, hourly_rate, sort_order) VALUES (?,?,?,?,?)',
            (pid, name, disc, rate, sort)
        )
        resource_id_map[(name, disc)] = cur.lastrowid
    con.commit()
    print(f'  Inserted {len(resources_design)} team resources')

    # ── 3. DRAWINGS ─────────────────────────────────────────────────────────
    print('Inserting drawings…')
    ws_del = wb['Deliverables Schedule ']
    del_rows = rows(ws_del)
    cur.execute('DELETE FROM drawings WHERE project_id=?', (pid,))

    current_disc = 'Architecture'
    current_series = None
    drawing_sort = 0
    drawing_count = 0
    seen_drawing_numbers = set()

    for i, r in enumerate(del_rows):
        # Discipline header (e.g. row[1] = 'ARCHITECTURE')
        if r[1] is not None and isinstance(r[1], str) and r[2] is None and r[3] is None:
            nd = norm_disc(r[1])
            if nd:
                current_disc = nd
                current_series = None
                continue

        # Series header (e.g. '00 SERIES DRAWINGS - GENERAL')
        if r[1] is not None and isinstance(r[1], str) and 'SERIES' in str(r[1]).upper():
            current_series = tx(r[1])
            continue

        # Drawing data row: r[1]==project_number(10162), r[2]=drawing_number
        if r[1] == 10162 and r[2] is not None:
            raw_num = str(r[2]).strip()
            # Deduplicate drawing numbers (schema enforces UNIQUE per project)
            if raw_num in seen_drawing_numbers:
                raw_num = raw_num + '-B'
            seen_drawing_numbers.add(raw_num)
            title = tx(r[3]) or 'Untitled'
            scale_val = tx(r[4]) if r[4] is not None else None
            drawing_sort += 1
            cur.execute(
                '''INSERT INTO drawings
                   (project_id, discipline, series, drawing_number, drawing_title,
                    scale, sort_order)
                   VALUES (?,?,?,?,?,?,?)''',
                (pid, current_disc, current_series, raw_num, title, scale_val, drawing_sort)
            )
            drawing_count += 1

    con.commit()
    print(f'  Inserted {drawing_count} drawings')

    # ── 4. C2C SNAPSHOTS ────────────────────────────────────────────────────
    print('Inserting C2C snapshots...')

    # Design phase snapshots
    design_sheets = [
        ('Cost to Complete - Start ', 1,  'design', '2025-09-08', 'Week 1 (Start)', 'old',  11),
        ('C2C Week 2',                2,  'design', '2025-09-15', 'Week 2',         'old',  10),
        ('C2C Week 3',                3,  'design', '2025-09-22', 'Week 3',         'new',   9),
        ('C2C Week 4',                4,  'design', '2025-09-29', 'Week 4',         'new',   8),
        ('C2C Week 5',                5,  'design', '2025-10-06', 'Week 5',         'new',   7),
        ('C2C Week 6',                6,  'design', '2025-10-13', 'Week 6',         'new',   6),
        ('C2C Week 9',                9,  'design', '2025-11-03', 'Week 9',         'new',   3),
        ('C2C Week 10',               10, 'design', '2025-11-10', 'Week 10',        'new',   2),
        ('C2C Week 11',               11, 'design', '2025-11-17', 'Week 11',        'new',   1),
        ('C2C Week 12',               12, 'design', '2025-11-24', 'Week 12',        'new',   1),
    ]
    # Construction phase snapshots
    cs_dates = {
        1: '2025-12-08', 2: '2025-12-15', 3: '2026-01-05',
        4: '2026-01-12', 5: '2026-01-19', 6: '2026-01-26',
        10: '2026-02-23', 11: '2026-03-02', 12: '2026-03-09',
    }
    cs_sheets = [(f'CS C2C Week {n}', n, 'construction', cs_dates[n], f'CS Week {n}', 'cs_old', 22-n) for n in sorted(cs_dates)]

    all_snapshots = design_sheets + cs_sheets

    for sheet_name, wk_num, phase, snap_date, wk_label, fmt, rem_weeks in all_snapshots:
        if sheet_name not in wb.sheetnames:
            print(f'  Skipping missing sheet: {sheet_name}')
            continue

        ws = wb[sheet_name]
        rs = rows(ws)

        # Insert snapshot
        cur.execute(
            '''INSERT INTO c2c_snapshots (project_id, phase, week_number, snapshot_date,
               week_label, snapshot_locked)
               VALUES (?,?,?,?,?,1)''',
            (pid, phase, wk_num, snap_date, wk_label)
        )
        snap_id = cur.lastrowid

        if fmt in ('old', 'cs_old'):
            _parse_old_format(cur, rs, snap_id, pid, resource_id_map, rem_weeks, fmt)
        else:
            _parse_new_format(cur, rs, snap_id, pid, resource_id_map, rem_weeks)

    con.commit()
    snap_count = cur.execute('SELECT COUNT(*) FROM c2c_snapshots WHERE project_id=?', (pid,)).fetchone()[0]
    alloc_count = cur.execute(
        'SELECT COUNT(*) FROM c2c_resource_allocations ra '
        'JOIN c2c_snapshots s ON s.id=ra.snapshot_id WHERE s.project_id=?', (pid,)
    ).fetchone()[0]
    fin_count = cur.execute(
        'SELECT COUNT(*) FROM c2c_discipline_financials df '
        'JOIN c2c_snapshots s ON s.id=df.snapshot_id WHERE s.project_id=?', (pid,)
    ).fetchone()[0]
    print(f'  Inserted {snap_count} snapshots, {alloc_count} resource allocations, {fin_count} discipline financials')

    # ── 5. APPROVALS ────────────────────────────────────────────────────────
    print('Inserting approvals…')
    cur.execute('DELETE FROM approvals WHERE project_id=?', (pid,))
    ws_app = wb['Approvals Tracker']
    app_rows = rows(ws_app)
    approval_data = [
        # (item_number, description, legislation, authority, category)
        ('1',    'Development Approval',               'Planning Act',                    None,            'Development'),
        ('2',    'Operational Works',                  'Planning Act',                    None,            'Operational Works'),
        ('2.01', 'Earthworks',                         None,                              None,            'Operational Works'),
        ('2.02', 'Stormwater',                         None,                              None,            'Operational Works'),
        ('3',    'Building Approval',                  None,                              None,            'Building'),
        ('3.01', 'QFES Approval',                      'Building Act',                    'QFES',          'Building'),
        ('3.02', 'QLEAVE',                             'Building Act / BCI',              'PLSL Authority','Building'),
        ('3.03', 'Building Approval',                  'Building Act / Building Regulation / NCC / QDC', 'Certifier', 'Building'),
        ('3.04', 'Build Over Sewer Approval',          'Queensland Development Code',     None,            'Building'),
        ('3.05', 'Build Over Stormwater Approval',     'Queensland Development Code',     None,            'Building'),
        ('3.06', 'Certificate of Occupancy',           'Building Act',                    'Certifier',     'Building'),
        ('4',    'Plumbing Approval',                  None,                              None,            'Plumbing'),
        ('4.01', 'Plumbing Approval',                  'Plumbing and Drainage Act',       None,            'Plumbing'),
        ('5',    'UU / UW Water & Sewer Connections',  'Water Supply Act',                'Unity Water',   'Utilities'),
        ('6',    'NBN',                                None,                              'NBN Co',        'Utilities'),
        ('7',    'Telstra',                            None,                              'Telstra',       'Utilities'),
        ('8',    'ENERGEX',                            None,                              'Energex',       'Utilities'),
    ]
    for sort, (item, desc, leg, auth, cat) in enumerate(approval_data):
        cur.execute(
            '''INSERT INTO approvals (project_id, item_number, description, legislation,
               authority, category, sort_order)
               VALUES (?,?,?,?,?,?,?)''',
            (pid, item, desc, leg, auth, cat, sort)
        )
    con.commit()
    print(f'  Inserted {len(approval_data)} approvals')

    # ── 6. CRITICAL ITEMS ───────────────────────────────────────────────────
    print('Inserting critical items…')
    cur.execute('DELETE FROM critical_items WHERE project_id=?', (pid,))
    critical_items = [
        # (item_number, details, agreed_strategy, action_step, responsible, action_date, date_raised, date_resolved, status, deliverable, initiator_group)
        ('1.1',  'Certainty of Project Cost',
                 'Be aware of cost implications of design', None, None, None, None, None, 'OPEN', None, 'Prekaro Projects'),

        ('2',    'Fillets to office tilt to match express joint',
                 'Replace tilt panel with lightweight construction with CFC cladding', 'Update drawings', 'Sonya', None, '2025-09-24', '2025-09-29', 'CLOSED', None, 'Prekaro - Architectural'),

        ('4',    'Entry soffit to office (CT6 to be painted concrete)',
                 'Add express jointed FC', 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('7',    'Wet areas to be 10mm wet area plasterboard',
                 None, 'Capture in wall types', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('8',    'Slab projection for lower awning under window of office',
                 'Revert to aluminium', 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('10',   'Hydrant booster cabinet to be located to the middle of the site between the crossovers',
                 None, 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('13',   'Mailbox wall/signage wall to be relocated next to pedestrian access',
                 None, 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('19',   'Palisade fencing 1800mm high included to street frontage only',
                 None, 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('20',   'Chainlink fencing 1800mm high with 3 strands of barbed wire to both side boundaries only',
                 None, 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('21',   'All roofing including office to be 0.42 trimdeck zinc',
                 None, 'Update schedule', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('22',   'Remove louvreclad spec. Sun blades 45mm x 450mm wide aluminium fixed at top & bottom of windows',
                 None, 'Update schedule', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('23',   'Board Room & Meeting Room - see if additional glass panels can be added to wall in lieu of full length plasterboard',
                 None, None, None, None, None, None, 'OPEN', None, 'Prekaro - Architectural'),

        ('24',   'Airlock Door to bathroom - can this be relocated outside Store Room?',
                 None, None, None, None, None, None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('25',   'All offices (1-5) - Prekaro to get preliminary pricing to amend to glazed suites',
                 None, None, None, None, None, None, 'OPEN', None, 'Prekaro - Architectural'),

        ('26',   'Delete L1 window from Kitchen into Warehouse',
                 None, None, None, None, None, None, 'CLOSED', None, 'Prekaro - Architectural'),

        ('1',    'Removal of panel plates and use of wind beam to tie together panels',
                 'Only referring to panel to panel connection plates & fire ties', 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Structural'),

        ('3',    'No steel beams to mezzanine slab - preferable for concrete',
                 'Please keep steel beam with Condeck', 'Instruction received from Sonya confirming steel beam with soffits is acceptable', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Structural'),

        ('5',    'Office roof structure outstanding',
                 None, 'Update drawings', None, None, '2025-09-24', None, 'CLOSED', None, 'Prekaro - Structural'),

        ('18',   'Drop edge beams to be incorporated into pavements in lieu of retaining walls to retaining areas 400mm and under',
                 None, 'Structural footing details to be updated', None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Structural'),

        ('23',   '20(c) Prepare Erosion and Sediment Control Plan',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('24',   'Arrange pre-start meeting with council for erosion control',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('26',   '24(a) Prepare Earthworks Drawings',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('27',   'On Site Drainage - Minor: stormwater connection and 114m3 detention storage required',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('28',   'Permanent Driveway Crossover - 9m wide Type B2 and 7.5m wide Type B2',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('29',   'Pavement design for footpath to be added',
                 None, None, None, None, '2025-09-24', None, 'OPEN', None, 'Prekaro - Civil'),

        ('4.1',  'Volume of warehouse and office exceeding allowable maximum volume requiring fire separation',
                 'Seek assessment of volume to u/s of structure', 'Advise certifier of volume to u/s structure', 'Elliot', '2025-09-10', '2025-09-10', '2025-09-11', 'CLOSED', '-', 'McVeigh - Architectural'),

        ('4.2',  'Timber stair structure - exposed steel stringers or concealed below the treads?',
                 'Timber stringer, engineering by stair manufacturer', 'Refer to Prekaro standard detail', None, None, '2025-09-12', '2025-09-23', 'CLOSED', '-', 'McVeigh - Architectural'),

        ('4.3',  'Office window sunshade - lower horizontal - aluminium not extension of slab?',
                 'Aluminium - amend to previous design', 'Amend to previous design', None, None, '2025-09-12', None, 'CLOSED', None, 'McVeigh - Architectural'),

        ('4.15', 'Padmount transformer coordination required',
                 None, None, None, '2025-11-10', None, None, 'OPEN', None, 'McVeigh - Architectural'),

        ('4.16', 'Required: Hydraulic drawings final issue',
                 None, None, None, '2025-11-10', None, None, 'OPEN', None, 'McVeigh - Architectural'),
    ]
    for it in critical_items:
        item_number, details, strategy, action_step, responsible, action_date, date_raised, date_resolved, status, deliverable, initiator = it
        cur.execute(
            '''INSERT INTO critical_items
               (project_id, item_number, details, agreed_strategy, action_step,
                responsible_person, action_date, date_raised, date_resolved, status,
                deliverable_affected, initiator_group)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, item_number, details, strategy, action_step, responsible,
             action_date, date_raised, date_resolved, status, deliverable, initiator)
        )
    con.commit()
    print(f'  Inserted {len(critical_items)} critical items')

    # ── 7. DESIGN CHANGES ───────────────────────────────────────────────────
    print('Inserting design changes…')
    cur.execute('DELETE FROM design_changes WHERE project_id=?', (pid,))
    design_changes = [
        ('1.01', '2025-09-29', 'Design Change', 'Robert Walsh', 'Architecture',
         'Please keep steel beam with Condeck. Add soffit to front entry to hide beam & Condeck. Projection to remain as aluminium hood.',
         'Client request', 'Office Level 1', 'Email from Robert Walsh dated 29.09.25',
         None, 'Yet to be submitted', 0, 0, None,
         0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 'Prekaro'),
        ('1.02', '2025-09-29', 'Design Change', 'Robert Walsh', 'Architecture',
         'Client wants to keep express joint cladding look to the office. Revert back to lightweight cladding with steel or concrete tilt columns.',
         'Client request', 'Office East elevation', 'Email from Robert Walsh dated 29.09.25',
         None, 'Yet to be submitted', 0, 0, None,
         0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 'Prekaro'),
    ]
    for dc in design_changes:
        (item_number, date_req, change_type, initiator_name, discipline,
         change_details, reason, area, doc_ref, var_ref, status,
         client_cost, risk_change, client_comments,
         arch, struc, civil, hyd, cert, lscape, fire_eng, fire_svc, builder, initiator_group) = dc
        cur.execute(
            '''INSERT INTO design_changes
               (project_id, item_number, date_requested, change_type, initiator_name,
                discipline, change_details, reason, area_location, document_reference,
                variation_reference, status, client_cost_impact, risk_assessment_change,
                client_comments, arch_fees, struc_fees, civil_fees, hyd_fees,
                certifier_fees, lscape_fees, fire_eng_fees, fire_services_fees,
                builder_dm_fees, initiator_group)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, item_number, date_req, change_type, initiator_name, discipline,
             change_details, reason, area, doc_ref, var_ref, status,
             client_cost, risk_change, client_comments,
             arch, struc, civil, hyd, cert, lscape, fire_eng, fire_svc, builder, initiator_group)
        )
    con.commit()
    print(f'  Inserted {len(design_changes)} design changes')

    # ── 8. LESSONS LEARNT ───────────────────────────────────────────────────
    print('Inserting lessons learnt…')
    cur.execute('DELETE FROM lessons_learnt WHERE project_id=?', (pid,))
    lessons = [
        ('1', None,
         'Structural steel to the office ground floor (supporting the slab) required 4 hour fire rating due to Type A construction',
         'Steel needed to be wrapped or sprayed with 60mm thickness. Promat advice obtained late required changing steel member thicknesses to heavier sections. Partition sizes needed to increase.',
         'Single fire compartment increased fire requirements',
         'Advice from certifier & BCA report', 0,
         'Identify Type A construction requirements early and get advice from supplier before finalising design',
         None, 'Update checklists to include check for fire requirements for steel',
         'Sonya Butt', '2025-12-17', None, 'Open', None),
    ]
    for ls in lessons:
        (item_number, client_ref, event_details, effect, cause, early_warnings,
         prev_identified, future_rec, action_step_ref, action_details,
         logged_by, logged_date, priority, status, responsible) = ls
        cur.execute(
            '''INSERT INTO lessons_learnt
               (project_id, item_number, event_details, effect, cause, early_warnings,
                previously_identified, future_recommendation, action_step_ref,
                action_details, logged_by, logged_date, priority, status, responsible_person)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, item_number, event_details, effect, cause, early_warnings,
             prev_identified, future_rec, action_step_ref, action_details,
             logged_by, logged_date, priority, status, responsible)
        )
    con.commit()
    print(f'  Inserted {len(lessons)} lessons learnt')

    # ── 9. SiD HAZARDS ──────────────────────────────────────────────────────
    print('Inserting SiD hazards…')
    cur.execute('DELETE FROM sid_hazards WHERE project_id=?', (pid,))
    ws_sid = wb['SiD Register']
    sid_rows_raw = rows(ws_sid)
    sid_hazards = _parse_sid(sid_rows_raw)
    for h in sid_hazards:
        cur.execute(
            '''INSERT INTO sid_hazards
               (project_id, ref_number, element_activity, hazard, potential_harm,
                likelihood, outcome, risk_rating, action_required, action_by,
                status, architect_notes, category)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid,) + h
        )
    con.commit()
    print(f'  Inserted {len(sid_hazards)} SiD hazards')

    # ── 10. VALUE LOG ────────────────────────────────────────────────────────
    print('Inserting value log…')
    cur.execute('DELETE FROM value_log WHERE project_id=?', (pid,))
    value_entries = [
        (None, None, '1',  'Value engineering item (structural)',                         '2005-07-25', 'PG',  'Structural',    None,  None,        'Email', 0),
        (None, None, '2a', 'No fire rating requirement between office and warehouse',     '2025-09-11', 'EB',  'Architecture',  None,  '2025-09-10', 'Email', 0),
        (None, None, '2b', 'Cost reductions in doors and windows as no FRL required',    '2025-09-11', 'EB',  'Architecture',  None,  '2025-09-10', 'Email', 0),
        (None, None, '2c', 'Potential for future expansion',                             '2025-09-11', 'EB',  'Architecture',  None,  '2025-09-10', 'Email', 0),
        (None, None, '2d', 'No wall wetting required',                                   '2025-09-11', 'EB',  'Architecture',  None,  '2025-09-10', 'Email', 0),
        (None, None, '2e', 'Simpler construction and detailing',                         '2025-09-11', 'EB',  'Architecture',  None,  '2025-09-10', 'Email', 0),
    ]
    for file_ref, job_ref, item_number, description, date, who, team, value_amount, comm_date, comm_how, approved in value_entries:
        cur.execute(
            '''INSERT INTO value_log
               (project_id, file_ref, job_ref, item_number, description, date,
                who, team, value_amount, communicated_date, communicated_how, approved)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, file_ref, job_ref, item_number, description, date,
             who, team, value_amount, comm_date, comm_how, approved)
        )
    con.commit()
    print(f'  Inserted {len(value_entries)} value log entries')

    # ── 11. BRIEF COMPLIANCE ─────────────────────────────────────────────────
    print('Inserting brief compliance…')
    cur.execute('DELETE FROM brief_compliance WHERE project_id=?', (pid,))
    ws_bc = wb['Brief Compliance Register']
    bc_rows = rows(ws_bc)
    bc_count = _parse_brief_compliance(cur, bc_rows, pid)
    con.commit()
    print(f'  Inserted {bc_count} brief compliance items')

    # ── Done ─────────────────────────────────────────────────────────────────
    con.close()
    print('\nDone. Demo project seeded successfully.')
    print(f'  Project ID: {pid}')
    print(f'  Project number: 10162 — Lot 31 The Hub Heathwood')


# ─── C2C parsers ─────────────────────────────────────────────────────────────

VALID_LIKELIHOOD = {'Almost Certain', 'Likely', 'Possible', 'Unlikely', 'Very Unlikely'}
VALID_OUTCOME    = {'Catastrophic', 'Major', 'Moderate', 'Minor', 'Insignificant'}
VALID_RISK       = {'Extreme', 'Significant', 'Moderate', 'Low', 'Negligible'}

def _safe_likelihood(v):
    s = tx(v)
    if s and s.title() in VALID_LIKELIHOOD:
        return s.title()
    if s and s in VALID_LIKELIHOOD:
        return s
    return None

def _safe_outcome(v):
    s = tx(v)
    if s and s.title() in VALID_OUTCOME:
        return s.title()
    if s and s in VALID_OUTCOME:
        return s
    return None

def _safe_risk(v):
    s = tx(v)
    if s and s.title() in VALID_RISK:
        return s.title()
    if s and s in VALID_RISK:
        return s
    return None


def _parse_old_format(cur, rs, snap_id, pid, resource_id_map, rem_weeks, fmt):
    """
    Parse old C2C format (Week 1, Week 2, CS weeks).
    Disciplines: Architecture (rows 30-39), Civil (41-49), Structural (51-60)
    0-indexed rows: 29-38, 40-48, 50-59
    """
    sections = [
        ('Architecture', 29, 31, 35, 36, 37, 38),   # disc, header_row, res_start, res_end, ctc, fee_wip, residual
        ('Civil',        40, 42, 44, 45, 46, 47),
        ('Structural',   50, 52, 55, 56, 57, 58),
    ]

    if fmt == 'cs_old':
        # CS uses wider columns (21 weeks) - util cols at index 2 to 2+21-1=22, hours at 23, amount at 24
        util_end_col = 23  # exclusive
        hours_col = 23
        amount_col = 24
    else:
        # Design old: util cols at index 2 to 12, hours at 12, amount at 13
        util_end_col = 12
        hours_col = 12
        amount_col = 13

    for disc, hdr, res_start, res_end, ctc_row, fee_row, residual_row in sections:
        resources_in_disc = []
        for ri in range(res_start, res_end + 1):
            if ri >= len(rs):
                break
            r = rs[ri]
            name = tx(r[0]) if r[0] is not None else None
            rate = f(r[1])
            if not name or name in ('Name', 'RESOURCE') or rate == 0:
                continue
            # Sum utilisations across all future weeks
            utils = [f(r[c]) for c in range(2, util_end_col) if c < len(r)]
            total_util = sum(utils)
            total_hours = round(total_util * 37.5, 4)
            cost = f(r[amount_col]) if amount_col < len(r) else total_hours * rate
            if cost == 0 and total_hours > 0:
                cost = total_hours * rate
            resources_in_disc.append((name, disc, total_util, rem_weeks, total_hours, cost))

        # Insert resource allocations
        for name, disc, total_util, rw, hours, cost in resources_in_disc:
            res_id = resource_id_map.get((name, disc))
            if res_id is None:
                # Try to find by name in same discipline
                res_id = _find_or_skip_resource(cur, pid, name, disc, rate, resource_id_map)
            if res_id is None:
                continue
            weekly_util = min(total_util / rw, 1.0) if rw > 0 else 0.0
            cur.execute(
                '''INSERT OR IGNORE INTO c2c_resource_allocations
                   (snapshot_id, resource_id, weekly_utilisation, remaining_weeks, cost_calculated)
                   VALUES (?,?,?,?,?)''',
                (snap_id, res_id, round(weekly_util, 6), rw, round(cost, 2))
            )

        # Insert discipline financials
        ctc = f(rs[ctc_row][amount_col]) if ctc_row < len(rs) and amount_col < len(rs[ctc_row]) else 0.0
        fee_wip = f(rs[fee_row][amount_col]) if fee_row < len(rs) and amount_col < len(rs[fee_row]) else 0.0
        # For old format: synergy_net_residual = fee_wip, total_net_to_carry = 0
        # adjusted_net_residual (generated) = fee_wip - 0 = fee_wip
        # under_over (generated) = fee_wip - ctc
        cur.execute(
            '''INSERT OR IGNORE INTO c2c_discipline_financials
               (snapshot_id, discipline, agreed_fee, cost_at_close, net_to_carry,
                synergy_net_residual, total_net_to_carry, construction_doc_cost_to_complete)
               VALUES (?,?,?,?,?,?,?,?)''',
            (snap_id, disc, 0.0, 0.0, 0.0, round(fee_wip, 2), 0.0, round(ctc, 2))
        )


def _find_or_skip_resource(cur, pid, name, disc, rate, resource_id_map):
    """Try to look up a resource by name and discipline, creating if needed."""
    row = cur.execute(
        'SELECT id FROM team_resources WHERE project_id=? AND name=? AND discipline=?',
        (pid, name, disc)
    ).fetchone()
    if row:
        resource_id_map[(name, disc)] = row[0]
        return row[0]
    # Create new resource
    cur.execute(
        'INSERT INTO team_resources (project_id, name, discipline, hourly_rate, sort_order) VALUES (?,?,?,?,99)',
        (pid, name, disc, rate)
    )
    new_id = cur.lastrowid
    resource_id_map[(name, disc)] = new_id
    return new_id


def _parse_new_format(cur, rs, snap_id, pid, resource_id_map, rem_weeks):
    """
    Parse new C2C format (Week 3–12).
    Layout (0-indexed rows):
      Row 9  (index 9):  TEAM RESOURCING header, week commencing dates at cols 8-16
      Row 10 (index 10): discipline label at col 1
      Row 11 (index 11): RESOURCE / RATE header
      Rows 12-17:        resources (name@1, rate@2, utils@8-16, hours@17, cost@18)
      Row 18 (index 18): Agreed Fee summary at col 3; Const Doc Cost@17
      Row 19 (index 19): Cost at close; Synergy Net Residual@17
      Row 20 (index 20): Net to Carry; Total Net to Carry@17
      Row 21 (index 21): Adjusted Net Residual@17
      Row 22 (index 22): Under/Over@17

    Discipline sections repeat every ~14 rows.
    """
    UTIL_COLS  = list(range(8, 17))   # indices 8-16 = weeks 3-11
    HOURS_COL  = 17
    COST_COL   = 18
    FINANCIAL_LABELS = {
        'agreed fee':                               'agreed_fee',
        'cost at close':                            'cost_at_close',
        'net to carry':                             'net_to_carry',
        'construction documentation - cost to complete': 'ctc',
        'synergy net residual':                     'synergy',
        'total net to carry':                       'total_ntc',
        'adjusetd net residual':                    'adj_residual',  # note typo in Excel
        'adjusted net residual':                    'adj_residual',
        'under/over':                               'under_over',
    }

    # Find all discipline section start rows by scanning col 1 for known disciplines
    disc_sections = []
    for i, r in enumerate(rs):
        if r[1] is None:
            continue
        nd = norm_disc(str(r[1]))
        if nd and (i == 0 or not norm_disc(str(rs[i-1][1])) if rs[i-1][1] else True):
            # Check it's a section header (col 2 has 'RATE' or next row has 'RESOURCE')
            disc_sections.append((i, nd))

    # Process each section
    for sec_idx, (sec_row, disc) in enumerate(disc_sections):
        end_row = disc_sections[sec_idx + 1][0] if sec_idx + 1 < len(disc_sections) else len(rs)

        resources_in_disc = []
        financials = {}

        for ri in range(sec_row, end_row):
            r = rs[ri]
            # Resource row: index 1 = name (str), index 2 = rate (numeric)
            name = tx(r[1]) if len(r) > 1 else None
            rate_val = r[2] if len(r) > 2 else None
            if (name and isinstance(rate_val, (int, float)) and rate_val > 0
                    and name not in ('RESOURCE', 'ARCHITECTURE', 'CIVIL', 'STRUCTURES',
                                     'HYDRAULICS', 'LANDSCAPING', 'CERTIFIER',
                                     'FIRE ENGINEERING', 'FIRE SERVICES')):
                utils = [f(r[c]) for c in UTIL_COLS if c < len(r)]
                total_util = sum(utils)
                hours = f(r[HOURS_COL]) if HOURS_COL < len(r) else total_util * 37.5
                cost = f(r[COST_COL]) if COST_COL < len(r) else hours * f(rate_val)
                if cost == 0 and hours > 0:
                    cost = hours * f(rate_val)
                resources_in_disc.append((name, disc, total_util, hours, cost, f(rate_val)))

            # Financial summary rows: label at index 3 or 17
            for label_col, val_col in ((3, 4), (3, 17), (17, 18)):
                if len(r) > val_col:
                    lbl = tx(r[label_col])
                    if lbl and lbl.lower() in FINANCIAL_LABELS:
                        key = FINANCIAL_LABELS[lbl.lower()]
                        if key not in financials:
                            financials[key] = f(r[val_col])

        # Insert resource allocations
        for name, disc_r, total_util, hours, cost, rate_val in resources_in_disc:
            res_id = resource_id_map.get((name, disc_r))
            if res_id is None:
                res_id = _find_or_skip_resource(cur, pid, name, disc_r, rate_val, resource_id_map)
            if res_id is None:
                continue
            weekly_util = min(total_util / rem_weeks, 1.0) if rem_weeks > 0 else 0.0
            cur.execute(
                '''INSERT OR IGNORE INTO c2c_resource_allocations
                   (snapshot_id, resource_id, weekly_utilisation, remaining_weeks, cost_calculated)
                   VALUES (?,?,?,?,?)''',
                (snap_id, res_id, round(weekly_util, 6), rem_weeks, round(cost, 2))
            )

        # Insert discipline financials
        # DB formula: adjusted = synergy - total_ntc; under_over = adjusted - ctc
        # Excel: total_ntc is negative when over budget, so DB needs positive to subtract correctly
        synergy  = financials.get('synergy', 0.0)
        excel_total_ntc = financials.get('total_ntc', 0.0)
        # Negate: if excel is -5297 (overrun), store 5297 so formula: 38586 - 5297 = 33289
        db_total_ntc = -excel_total_ntc

        # agreed_fee: sum of all phase values shown at cols 4,5,6,7
        agreed_fee = 0.0
        for ri in range(sec_row, end_row):
            r = rs[ri]
            if len(r) > 3 and tx(r[3]) and 'agreed fee' in str(r[3]).lower():
                for c in range(4, 8):
                    if c < len(r):
                        agreed_fee += f(r[c])
                break

        cur.execute(
            '''INSERT OR IGNORE INTO c2c_discipline_financials
               (snapshot_id, discipline, agreed_fee, cost_at_close, net_to_carry,
                synergy_net_residual, total_net_to_carry, construction_doc_cost_to_complete)
               VALUES (?,?,?,?,?,?,?,?)''',
            (snap_id, disc,
             round(agreed_fee, 2),
             round(financials.get('cost_at_close', 0.0), 2),
             round(financials.get('net_to_carry', 0.0), 2),
             round(synergy, 2),
             round(db_total_ntc, 2),
             round(financials.get('ctc', 0.0), 2))
        )


def _parse_sid(rs):
    """Parse SiD Register rows into tuples for insertion."""
    hazards = []
    VALID_LIKELIHOODS = {'Almost Certain', 'Likely', 'Possible', 'Unlikely', 'Very Unlikely'}
    VALID_OUTCOMES    = {'Catastrophic', 'Major', 'Moderate', 'Minor', 'Insignificant'}
    VALID_RISKS       = {'Extreme', 'Significant', 'Moderate', 'Low', 'Negligible'}
    current_category  = None
    SKIP_PREFIXES = ('PROJECT', 'CLIENT', 'REVIEW', 'MCVEIGH', 'STAKEHOLDER',
                     'REF', 'ITEM', 'ELEMENT', 'HAZARD', 'POTENTIAL')

    for r in rs:
        if not any(c is not None for c in r):
            continue

        ref = tx(r[0])
        hazard_text = tx(r[2])
        if ref is None or hazard_text is None:
            continue
        # Skip project header rows and column header rows
        if any(str(ref).upper().startswith(p) for p in SKIP_PREFIXES):
            continue
        if 'hazard' in str(hazard_text).lower() and len(hazard_text) > 60:
            continue

        # Category section headers (no hazard details present)
        if r[1] is None and r[3] is None and r[4] is None:
            current_category = ref
            continue

        element   = tx(r[1])
        harm      = tx(r[3])
        raw_lk    = tx(r[4])
        raw_out   = tx(r[5])
        raw_risk  = tx(r[6])
        action    = tx(r[7])
        action_by = tx(r[8])
        status    = tx(r[9]) or 'Open'
        notes     = tx(r[10])

        likelihood = raw_lk.title() if raw_lk and raw_lk.title() in VALID_LIKELIHOODS else None
        outcome    = raw_out.title() if raw_out and raw_out.title() in VALID_OUTCOMES else None
        risk_rating = raw_risk.title() if raw_risk and raw_risk.title() in VALID_RISKS else None

        hazards.append((
            str(ref), element, hazard_text, harm,
            likelihood, outcome, risk_rating,
            action, action_by, status, notes, current_category
        ))
    return hazards


def _parse_brief_compliance(cur, rs, pid):
    """Parse Brief Compliance Register rows and insert."""
    count = 0
    source_doc = None
    data_started = False

    for r in rs:
        # Source document header
        if r[0] is not None and isinstance(r[0], str) and 'brief document' in str(r[0]).lower():
            source_doc = tx(r[0])
            continue
        # Column header row
        if r[0] is not None and str(r[0]).strip() in ('Spec. No.', 'Spec No.'):
            data_started = True
            continue
        if not data_started:
            continue

        spec_no = tx(r[0])
        if spec_no is None:
            # Could be a new source document header
            if r[3] is not None and 'brief document' in str(r[3]).lower().strip()[:20]:
                source_doc = tx(r[3])
            continue

        # New source doc mid-sheet
        if isinstance(r[0], str) and 'brief document' in str(r[0]).lower():
            source_doc = tx(r[0])
            continue

        location   = tx(r[1])
        clause     = tx(r[2])
        brief_item = tx(r[3])
        discipline = tx(r[4])
        compliant  = 1 if tx(r[5]) and str(r[5]).strip().upper() in ('Y', 'YES', '1') else 0
        deviation  = 1 if tx(r[6]) and str(r[6]).strip().upper() in ('Y', 'YES', '1') else 0
        comments   = tx(r[7])
        client_resp = tx(r[8])
        counter    = tx(r[9])

        if brief_item is None:
            continue

        cur.execute(
            '''INSERT INTO brief_compliance
               (project_id, spec_number, location, clause, brief_item, discipline,
                compliant, deviation, comments, client_response, counter_response,
                source_document)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, spec_no, location, clause, brief_item, discipline,
             compliant, deviation, comments, client_resp, counter, source_doc)
        )
        count += 1
    return count


if __name__ == '__main__':
    main()
